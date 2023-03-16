import os
import time
from datetime import datetime

import serial
import pyvisa as visa
from io import StringIO

import tkinter as tk
import tkinter.messagebox as msgbox
from tkinter import filedialog

import img2pdf

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import builtin_format_code

import pandas as pd
import numpy as np

import threading

from Band_list import *
import Tx_measure_Apt as Apt
import Tx_measure_Nonsig as Nonsig
import Tx_measure_Signaling as Sig

font_style = Font(
    name="Calibri",
    size=10,
    bold=False,
    italic=False,
    vertAlign=None,  # 첨자
    underline="none",  # 밑줄
    strike=False,  # 취소선
    color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
)


class Open_Dut:
    def __init__(self, Comport):
        self.ser = serial.Serial(
            port=Comport,
            baudrate=115200,
            bytesize=8,
            parity="N",
            stopbits=1,
            timeout=0.3,
            xonxoff=False,
            rtscts=False,
            dsrdtr=False,
        )

    def serial_open(self):
        self.ser.o()

    def serial_close(self):
        self.ser.c()

    def con_check(self):
        if self.ser.is_open == 0:
            return 0
        else:
            return 1

    def at_write(self, cmd):
        self.cmd_add = cmd + "\r\n"
        self.ser.reset_input_buffer()
        self.ser.write(self.cmd_add.encode())
        self.raw_responses = self.ser.readlines()
        self.ser.reset_output_buffer()

        self.responses = []
        for line_response in self.raw_responses:
            self.responses.append(line_response.strip().decode())
        # print(self.responses)
        return self.responses  # , self.read

    def check_bcm(self, text_area):
        while True:
            response = self.ser.readlines()
            for line_response in response:
                # 'utf-8' codec can't decode byte 0xff in position 0: invalid start byte : decode() -> decode('latin1') 으로 변경
                line = line_response.strip().decode("latin1")
                text_area.insert(tk.END, f"{line}\n")
                text_area.see(tk.END)
                if line == "BOOTING COMPLETED":
                    return line


def return_print(*prt_str):  # print 출력값을 변수로 리턴
    io = StringIO()
    print(*prt_str, file=io, sep=",", end="")
    return io.getvalue()


def createDirectory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print("\tError: Failed to create directory.")


def add_file(log_path):
    default_path = os.path.join(
        os.path.expanduser("~"), "C:\\DIST\\CONFIG\\DASEUL\\CABLE_LOSS\\"
    )
    log_path.delete(0, tk.END)
    files = filedialog.askopenfilename(
        title="loss 파일을 선택하세요",
        filetypes=(("dec 파일", "*.dec"), ("txt 파일", "*.txt"), ("모든 파일", "*.*")),
        initialdir=default_path,
    )
    # 사용자가 선택한 파일 목록
    log_path.insert(tk.END, files)
    Selected_path = log_path.get()

    if Selected_path == "":
        msgbox.showwarning("경고", "Loss 파일을 선택하세요")
        return


def resut_to_excel(result, file_name, name_of_sheet):
    if not os.path.exists(file_name):  # 파일이 없으면 쓰기모드
        with pd.ExcelWriter(file_name, mode="w", engine="openpyxl") as writer:
            result.to_excel(writer, sheet_name=name_of_sheet)

    else:  # 파일이 있으면 어펜드 모드
        with pd.ExcelWriter(file_name, mode="a", engine="openpyxl") as writer:
            result.to_excel(writer, sheet_name=name_of_sheet)


def WB_Format(filename, i, j, k):
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        for row_c in range(i, row_max + 1, 1):
            for col_c in range(j, col_max + 1, 1):
                wb[sheet].cell(row=row_c, column=col_c).font = font_style
                wb[sheet].cell(row=row_c, column=col_c).alignment = Alignment(
                    horizontal="center"
                )
                # wb[sheet].cell(row=row_c, column=col_c).number_format = "#,##0.0"
                wb[sheet].cell(
                    row=row_c, column=col_c
                ).number_format = builtin_format_code(k)
    wb.save(filename)


def Check_mipi_R(text_area, rat, mipi_data, combo4):
    text_area.delete("1.0", tk.END)
    Comport = combo4.get()
    ch = mipi_data[0].get()
    usid = mipi_data[1].get()
    addr = mipi_data[2].get()

    dut = Open_Dut(Comport)
    at_resp = dut.at_write("AT+MODECHAN=0,0")
    at_resp = dut.at_write("AT+HNSSTOP")
    if rat == "LTE":
        at_resp = dut.at_write("AT+LRFFINALSTART=1,1")
    elif rat == "NR":
        # AT+NRFFINALSTART=Band, SA
        # SA : 0, NSA : 1
        at_resp = dut.at_write("AT+NRFFINALSTART=1,0")
    text_area.insert(tk.END, f"{at_resp[0]:<25}\t|\t{at_resp[2:3:2]}\n")
    at_resp = dut.at_write(f"AT+MIPIREAD = {ch},{usid},{addr}")
    text_area.insert(tk.END, f"{at_resp[0]:<25}\t|\t{at_resp[2:3:2]}\n")
    at_resp = dut.at_write("AT+MODECHAN=0,2")
    text_area.see(tk.END)


def Check_mipi_W(text_area, rat, mipi_data, combo4):
    text_area.delete("1.0", tk.END)
    Comport = combo4.get()
    ch = mipi_data[0].get()
    usid = mipi_data[1].get()
    addr = mipi_data[2].get()
    data = mipi_data[3].get()

    dut = Open_Dut(Comport)
    at_resp = dut.at_write("AT+MODECHAN=0,0")
    at_resp = dut.at_write("AT+HNSSTOP")
    if rat == "LTE":
        at_resp = dut.at_write("AT+LRFFINALSTART=1,1")
    elif rat == "NR":
        # AT+NRFFINALSTART=Band, SA
        # SA : 0, NSA : 1
        at_resp = dut.at_write("AT+NRFFINALSTART=1,0")
    text_area.insert(tk.END, f"{at_resp[0]:<25}\t|\t{at_resp[2:3:2]}\n")
    at_resp = dut.at_write(f"AT+MIPIWRITE = {ch},{usid},{addr},{data}")
    text_area.insert(tk.END, f"{at_resp[0]:<25}\t|\t{at_resp[2:3:2]}\n")
    at_resp = dut.at_write("AT+MODECHAN=0,2")
    text_area.see(tk.END)


def Read_mipi(dut, rat, mipi_data, addr, text_area):
    ch = mipi_data[0].get()
    if rat == "LTE":
        usid = mipi_data[1].get()
        at_resp = dut.at_write(f"AT+MIPIREAD = {ch},{usid},{addr}")
    elif rat == "NR":
        usid = hex(int(mipi_data[1].get())).strip("0x")
        at_resp = dut.at_write(f"AT+NMIPIREAD = {ch},{usid},{addr}")

    return at_resp[2]


def Write_mipi(dut, rat, mipi_data, addr, data, text_area):
    ch = mipi_data[0].get()
    if rat == "LTE":
        usid = mipi_data[1].get()
        at_resp = dut.at_write(f"AT+MIPIWRITE = {ch},{usid},{addr}")
    elif rat == "NR":
        usid = hex(int(mipi_data[1].get())).strip("0x")
        at_resp = dut.at_write(f"AT+NMIPIWRITE = {ch},{usid},{addr}")
    print(f"Write Bias = {data}, {at_resp[2:]}")
    return at_resp[2]


def Hmipi_read(dut, mipi_data, addr, text_area):
    ch = mipi_data[0].get()
    usid = hex(int(mipi_data[1].get())).strip("0x")
    at_resp = dut.at_write(f"AT+HREADMIPI = {ch},{usid},{addr}")

    return at_resp


def Power_on(combo2, combo3, combo4, text_area):
    text_area.delete("1.0", tk.END)
    Sys_power = combo2.get()
    Pa_power = combo3.get()
    Comport = combo4.get()

    if Sys_power == "":
        msgbox.showwarning("Warning", "Check System Power")
        return
    elif Pa_power == "":
        msgbox.showwarning("Warning", "Check PA VCC")
        return

    rm = visa.ResourceManager()
    E3632a_1 = rm.open_resource(Sys_power)  # SYS전류측정 Power supply
    if Pa_power == "NO Supply for PA":
        E3632a_2 = "NO Supply for PA"
    else:
        E3632a_2 = rm.open_resource(Pa_power)  # PA전류측정 Power supply

    text_area.insert(tk.END, f"Start threading\n")
    text_area.insert(tk.END, f"Thread count = {threading.active_count()}\n")

    for thread in threading.enumerate():
        text_area.insert(
            tk.END,
            f"Thread ID = {threading.get_ident()}, Thread name = {thread.name}\n",
        )
    text_area.see(tk.END)

    text_area.insert(tk.END, f"\nSystem Power = {Sys_power}\t|\tPA VCC = {Pa_power}\n")
    text_area.see(tk.END)

    # OCP Check 후 Sys_Power Off -> On
    recycle = OCP_Check(E3632a_1, E3632a_2, text_area)
    Psupply_vcc_on_off(E3632a_1, 4.5)

    dut = Open_Dut(Comport)
    response = dut.check_bcm(text_area)

    if response == "BOOTING COMPLETED":
        msgbox.showinfo("Info", "BOOTING COMPLETED")


def Start(
    Main_loss_path,
    Sub_loss_path,
    combo1,
    combo2,
    combo3,
    combo4,
    Run_mode_var,
    Rat_option_var,
    Ch_option_var,
    User_defined_path,
    User_defined_band,
    User_defined_ch,
    Band_Select_Main_var,
    Band_Select_Sub_var,
    BW_list,
    Pw_option_var,
    Mipi_data,
    canvas,
    fig,
    ax1,
    ax2,
    ax3,
    text_area,
):
    try:
        Main_loss = Main_loss_path.get()
        Sub_loss = Sub_loss_path.get()
        text_area.delete("1.0", tk.END)
        Run_mode = Run_mode_var.get()
        Equip = combo1.get()
        Sys_power = combo2.get()
        Pa_power = combo3.get()
        Comport = combo4.get()

        if Equip == "TCPIP::127.0.0.1::INSTR":
            C_Box = "CMW100"
        elif Equip == "GPIB0::20::INSTR":
            C_Box = "CMW500"

        text_area.insert(tk.END, f"Start threading\n")
        text_area.insert(tk.END, f"Thread count = {threading.active_count()}\n")

        for thread in threading.enumerate():
            text_area.insert(
                tk.END,
                f"Thread ID = {threading.get_ident()}, Thread name = {thread.name}\n",
            )
        text_area.see(tk.END)

        if Equip == "":
            msgbox.showwarning("Warning", "Check Call Box")
            return
        elif Sys_power == "":
            msgbox.showwarning("Warning", "Check System Power")
            return
        elif Pa_power == "":
            msgbox.showwarning("Warning", "Check PA VCC")
            return
        elif Comport == "":
            msgbox.showwarning("Warning", "Check Comport")
            return
        elif (C_Box == "CMW500") & (Rat_option_var.get() == 3):
            msgbox.showwarning("Warning", "Choose CMW100")
            return

        rm = visa.ResourceManager()
        Callbox = rm.open_resource(Equip)
        Callbox.timeout = 20000
        E3632a_1 = rm.open_resource(Sys_power)  # SYS전류측정 Power supply
        if Pa_power == "NO Supply for PA":
            E3632a_2 = "NO Supply for PA"
        else:
            E3632a_2 = rm.open_resource(Pa_power)  # PA전류측정 Power supply

        text_area.insert(
            tk.END,
            f"\nCallBox = {Equip}\t|\tSystem Power = {Sys_power}\t|\tPA VCC = {Pa_power}\n",
        )
        text_area.see(tk.END)
        # Callbox reset 시 losstable setting 진행
        Sig.Callbox_reset(Main_loss, Equip, Callbox, text_area)

        Test_band_ch_list = Check_testband(
            Rat_option_var,
            Ch_option_var,
            User_defined_path,
            User_defined_band,
            User_defined_ch,
            Band_Select_Main_var,
            Band_Select_Sub_var,
        )
        pwr_levels = Check_pwr_lvs(Rat_option_var.get(), Pw_option_var.get())
        today = datetime.today().strftime("%Y_%m_%d")
        start_T = datetime.today().strftime("%Y%m%d_%H%M")
        save_dir = os.getcwd() + "\\Measurement_Result\\" + today + "\\"
        createDirectory(save_dir)

        dut = Open_Dut(Comport)
        recycle = OCP_Check(E3632a_1, E3632a_2, text_area)

        if recycle:
            bcm_check = dut.check_bcm(text_area)

        if not Test_band_ch_list:
            msgbox.showwarning("Warning", "Select Test Band")
            return

        # Singaling test
        elif Run_mode == 1:
            if Rat_option_var.get() == 1:  # 3G
                msgbox.showwarning("Warning", "Not supported yet")
                return

            elif Rat_option_var.get() == 2:  # LTE
                filename = save_dir + f"LTE_Current_{C_Box}_{start_T}.xlsx"
                Sig.LTE_Signaling_test(
                    dut,
                    Equip,
                    Callbox,
                    E3632a_1,
                    E3632a_2,
                    Main_loss,
                    Sub_loss,
                    "LTE",
                    Test_band_ch_list,
                    BW_list,
                    pwr_levels,
                    canvas,
                    fig,
                    ax1,
                    ax2,
                    ax3,
                    save_dir,
                    filename,
                    text_area,
                )

        # Non-Singaling test
        elif Run_mode == 2:
            if Rat_option_var.get() == 1:  # 3G
                msgbox.showwarning("Warning", "Not supported yet")
                return

            elif Rat_option_var.get() == 2:  # LTE
                filename = save_dir + f"LTE_NonSig_{C_Box}_{start_T}.xlsx"
                Callbox.write("SYST:DISP:UPD ON")
                Nonsig.Set_factolog(dut, text_area)

                for key in Test_band_ch_list:
                    if key == 0:
                        path = "Main"
                        loss_table(Main_loss, "Loss_table_Main", Equip, Callbox)
                    elif key == 1:
                        path = "Sub"
                        loss_table(Sub_loss, "Loss_table_Sub", Equip, Callbox)

                    for Testband in Test_band_ch_list[key]:
                        channel_list = Test_band_ch_list[key][Testband]
                        bandwidth_list = BW_list[key][Testband]
                        band_tx_result = []
                        bandwidth_tx_result = []
                        for bandwidth in bandwidth_list:
                            ch_tx_result = []
                            for channel in channel_list[bandwidth]:
                                tx_result = Nonsig.LTE_tx_measure(
                                    dut,
                                    Equip,
                                    Callbox,
                                    E3632a_1,
                                    E3632a_2,
                                    path,
                                    Testband,
                                    channel,
                                    bandwidth,
                                    pwr_levels,
                                    canvas,
                                    fig,
                                    ax1,
                                    ax2,
                                    ax3,
                                    save_dir,
                                    text_area,
                                )
                                Nonsig.End_testmode(dut, text_area)
                                ch_tx_result.append(tx_result)

                            bandwidth_tx_result = pd.concat(
                                ch_tx_result,
                                axis=1,
                                keys=[
                                    f"CH {i}"
                                    for i in range(1, len(channel_list[bandwidth]) + 1)
                                ],
                                # names=[f"LTE {path} B{Testband} {bandwidth}MHz"],
                            )
                            band_tx_result.append(bandwidth_tx_result)
                        Total_tx_result = pd.concat(
                            band_tx_result,
                            axis=0,
                            keys=[str(bw) + " MHz" for bw in bandwidth_list],
                            names=[f"LTE {path} B{Testband}"],
                        )
                        resut_to_excel(
                            Total_tx_result, filename, f"LTE {path} B{Testband}"
                        )
                        WB_Format(filename, 1, 1, 0)

                images2PdfFile(save_dir, filename)

            elif Rat_option_var.get() == 3:
                filename = save_dir + f"NR_NonSig_{C_Box}_{start_T}.xlsx"
                Callbox.write("SYST:DISP:UPD ON")
                Nonsig.Set_factolog(dut, text_area)

                for key in Test_band_ch_list:
                    if key == 0:
                        path = "Main"
                        loss_table(Main_loss, "Loss_table_Main", Equip, Callbox)
                    elif key == 1:
                        path = "Sub"
                        loss_table(Sub_loss, "Loss_table_Sub", Equip, Callbox)

                    for Testband in Test_band_ch_list[key]:
                        channel_list = Test_band_ch_list[key][Testband]
                        bandwidth_list = BW_list[key][Testband]
                        band_tx_result = []
                        bandwidth_tx_result = []
                        for bandwidth in bandwidth_list:
                            ch_tx_result = []
                            for channel in channel_list[bandwidth]:
                                tx_result = Nonsig.NR_tx_measure(
                                    dut,
                                    Equip,
                                    Callbox,
                                    E3632a_1,
                                    E3632a_2,
                                    path,
                                    Testband,
                                    channel,
                                    bandwidth,
                                    pwr_levels,
                                    canvas,
                                    fig,
                                    ax1,
                                    ax2,
                                    ax3,
                                    save_dir,
                                    text_area,
                                )
                                Nonsig.End_testmode(dut, text_area)
                                ch_tx_result.append(tx_result)

                            bandwidth_tx_result = pd.concat(
                                ch_tx_result,
                                axis=1,
                                keys=[
                                    f"CH {i}"
                                    for i in range(1, len(channel_list[bandwidth]) + 1)
                                ],
                                # names=[f"NR {path} n{Testband} {bandwidth}MHz"],
                            )
                            band_tx_result.append(bandwidth_tx_result)

                        Total_tx_result = pd.concat(
                            band_tx_result,
                            axis=0,
                            keys=[str(bw) + " MHz" for bw in bandwidth_list],
                            names=[f"NR {path} n{Testband}"],
                        )
                        resut_to_excel(
                            Total_tx_result, filename, f"NR {path} n{Testband}"
                        )
                        WB_Format(filename, 1, 1, 0)

                images2PdfFile(save_dir, filename)

        # APT Tuning
        elif Run_mode == 3:
            loss_table(Main_loss, "Loss_table_Main", Equip, Callbox)
            filename = save_dir + f"NR_APT_Tuning_{C_Box}_{start_T}.xlsx"
            Callbox.write("SYST:DISP:UPD ON")

            Nonsig.Set_factolog(dut, text_area)

            for key in Test_band_ch_list:
                if key == 0:
                    path = "Main"
                    loss_table(Main_loss, "Loss_table_Main", Equip, Callbox)
                elif key == 1:
                    path = "Sub"
                    loss_table(Sub_loss, "Loss_table_Sub", Equip, Callbox)
                for Testband in Test_band_ch_list[key]:
                    channel_list = Test_band_ch_list[key][Testband]
                    bandwidth_list = BW_list[key][Testband]
                    band_tx_result = []
                    bandwidth_tx_result = []
                    for bandwidth in bandwidth_list:
                        ch_tx_result = []
                        for channel in channel_list[bandwidth]:
                            tx_result = Apt.Dsp_off_tx_measure(
                                dut,
                                Equip,
                                Callbox,
                                E3632a_1,
                                E3632a_2,
                                path,
                                Testband,
                                channel,
                                bandwidth,
                                pwr_levels,
                                Mipi_data,
                                canvas,
                                fig,
                                ax1,
                                ax2,
                                ax3,
                                save_dir,
                                text_area,
                            )
                            Nonsig.End_testmode(dut, text_area)
                            ch_tx_result.append(tx_result)

                        bandwidth_tx_result = pd.concat(
                            ch_tx_result,
                            axis=0,
                            keys=[
                                f"CH {i}"
                                for i in range(1, len(channel_list[bandwidth]) + 1)
                            ],
                            # names=[f"NR n{Testband} {channel}CH"],
                        )
                        band_tx_result.append(bandwidth_tx_result)
                    Total_tx_result = pd.concat(
                        band_tx_result,
                        axis=0,
                        keys=[str(bw) + " MHz" for bw in bandwidth_list],
                        names=[f"NR {path} n{Testband}"],
                    )
                    resut_to_excel(Total_tx_result, filename, f"NR {path} n{Testband}")
                    WB_Format(filename, 4, 2, 0)

            images2PdfFile(save_dir, filename)

        else:
            msgbox.showwarning("Warning", "Not supported")
            return

        text_area.see(tk.END)
        text_area.insert(tk.END, f"\n\n")
        response = dut.at_write("AT+MODECHAN=0,2")
        text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
        response = dut.at_write("AT+DISPTEST=0,3")
        text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
        text_area.see(tk.END)

        Psupply_off(E3632a_1, E3632a_2)
        msgbox.showinfo("Message", "***** 작업완료 *****")
        # rm.close()

    except ValueError as e:
        if str(e) == "Unable to process empty list":
            raise ValueError("Unable to process empty list")
        else:
            raise e
    except Exception as e:
        msgbox.showwarning("Warning", e)


def loss_table(Selected_path, table_name, Equip, Callbox):
    with open(Selected_path, "r", encoding="utf-8") as loss_file:
        fname, ext = os.path.splitext(Selected_path)
        if ext == ".txt":
            # 0번~4번 행 값 스킵하고, 각 컬럼명을 index와 value로 지정
            df = pd.read_csv(
                loss_file, sep="=", names=["index", "value"], skiprows=[0, 1, 2, 3, 4]
            )
        else:
            df = pd.read_csv(
                loss_file,
                sep="=",
                names=["index", "value"],
                # skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            )
    loss_file.close()
    Start_index = df.index[df["index"].str.contains("Measured_CableLoss_01")].tolist()
    df = df.loc[Start_index[0] + 2 :].reset_index()
    df = df[["index", "value"]]
    df = df[["value"]].assign(g=df.index % 2)

    Equipment = Callbox.query("*IDN?").split(",")[1]

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        df_freq = df.loc[df["g"] == 0, "value"].astype(int) * 1000000.00
        df_loss = df.loc[df["g"] == 1, "value"].astype(float).abs()  # loss value
        df["value"] = df_freq.combine_first(df_loss)  # 2개 Series freq 기준으로 병합
        loss_list = list(df["value"])  # value 값 loss_list에 list로 저장
        loss = return_print(*loss_list)

        Callbox.write(f"CONF:FDC:DEAC:ALL")
        Callbox.write(f"CONF:BASE:FDC:CTAB:DEL:ALL")
        Callbox.write(f"CONF:BASE:FDC:CTAB:CRE '{table_name}', {loss}")
        Callbox.write(
            f"CONF:CMWS:FDC:ACT:TX R118,"
            f" '{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}'"
        )
        Callbox.write(
            f"CONF:CMWS:FDC:ACT:RX R118,"
            f" '{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}','{table_name}'"
        )

    else:
        df_freq = (
            df.loc[df["g"] == 0, "value"].astype(int).astype(str) + "e6"
        )  # Frequency 정수로 변환 후 문자열로 변환 & e6 추가
        df_loss = df.loc[df["g"] == 1, "value"].astype(float).abs()  # loss value
        df["value"] = df_freq.combine_first(df_loss)  # 2개 Series freq 기준으로 병합
        loss_list = list(df["value"])  # value 값 loss_list에 list로 저장
        loss = return_print(*loss_list)

        Callbox.write(f"CONF:FDC:DEAC RF1C")
        Callbox.write(f"CONF:BASE:FDC:CTAB:CRE '{table_name}',{loss}")
        Callbox.write(f"CONF:FDC:ACT RF1C, '{table_name}', RXTX")


def Callback_Sys_P(combo2):
    Sys_vcc = combo2.get()
    if Sys_vcc:
        print(f"System Power = {Sys_vcc}")


def Callback_PA_Vcc(combo3):
    Pa_vcc = combo3.get()
    if Pa_vcc:
        print(f"PA VCC = {Pa_vcc}")


def Callback_Comport(combo4):
    Comport = combo4.get()
    if Comport:
        print(f"Comport = {Comport}")


def OCP_Check(Sys_vcc, Pa_vcc, text_area):
    P_recycle = False
    E3632a = [Sys_vcc, Pa_vcc]

    for Psupply in E3632a:
        if Psupply != "NO Supply for PA":
            OCP_stat = Psupply.query("CURR:PROT:TRIP?")
            OVP_stat = Psupply.query("VOLT:PROT:TRIP?")

            if Psupply == Sys_vcc:
                Volt = 4.3
            elif Psupply == Pa_vcc:
                Volt = 4.0

            if (OCP_stat == "1\n") or (
                OVP_stat == "1\n"
            ):  # 리턴은 0 or 1이 아니고 \n 까지 붙어서 옴
                text_area.insert(tk.END, f"{Psupply} OCP/OVP TRIPPED")
                Psupply_init(Psupply)
                Psupply_vcc_on_off(Psupply, Volt)
                if Psupply == Sys_vcc:  # System 전원을 off 했을때만 P_recycle True 설정
                    P_recycle = True

            Outp_stat = Psupply.query("OUTP:STAT?").strip("\n")
            # print(f"{Psupply} Outp_stat ={Outp_stat}")
            if (Outp_stat == "0") or (
                int(float(Psupply.query("MEAS:CURR?")) * 1000) <= 5
            ):
                Psupply_vcc_on_off(Psupply, Volt)
                if Psupply == Sys_vcc:  # System 전원을 off 했을때만 P_recycle True 설정
                    P_recycle = True

    text_area.see(tk.END)

    return P_recycle


def Psupply_init(Psupply):
    Psupply.write("OUTP OFF")
    Psupply.write("CURR:PROT:CLE")
    time.sleep(0.5)
    Psupply.write("VOLT:PROT:CLE")
    time.sleep(0.5)


def Psupply_vcc_on_off(Psupply, vcc):
    Psupply.write(f"OUTP OFF")
    time.sleep(0.5)
    Psupply.write(f"VOLT {vcc}")
    Psupply.write(f"VOLT:PROT 15")
    Psupply.write(f"VOLT:PROT:STAT ON")
    Psupply.write(f"CURR:PROT 7.5")
    Psupply.write(f"CURR:PROT:STAT ON")
    Psupply.write(f"OUTP ON")
    # Psupply.control_ren(0)


def Psupply_off(Sys_vcc, Pa_vcc):
    time.sleep(0.5)
    Pa_vcc.write("OUTP OFF")
    # Pa_vcc.control_ren(0)
    time.sleep(0.5)
    Sys_vcc.write("OUTP OFF")
    # Sys_vcc.control_ren(0)


def Equipment_scan():
    # visa.log_to_screen()        # 장비로그 On
    rm = visa.ResourceManager()  # Resource Manager 생성
    Resource = rm.list_resources()  # Visa resource list
    CB_list = []
    PS_list1 = []
    PS_list2 = []
    CP_list = []

    for r in Resource:
        if r.startswith("ASRL"):
            com = r.replace("ASRL", "COM")
            com = com.split("::")[0]
            CP_list.append(com)

        if r.startswith("GPIB"):
            instrument = rm.open_resource(r, timeout=500)
            Equipment = instrument.query("*IDN?").split(",")
            # Power Supply list
            if Equipment[1] == "E3632A":
                PS_list1.append(r)
                PS_list2.append(r)
            elif Equipment[1] == "CMW":
                CB_list.append(r)

    rm = visa.ResourceManager("@py")  # pyvisa-py로 검색해야 TCPIP 잘 찾음
    Resource = rm.list_resources()  # Visa resource list
    # instr_list = ["TCPIP::127.0.0.1::5025::SOCKET"]
    cwm100 = "TCPIP::127.0.0.1::INSTR"

    for r in Resource:
        # Callbox list
        if r == cwm100:
            instrument = rm.open_resource(r, timeout=2000)
            Equipment = instrument.query("*IDN?").split(",")
            if Equipment[1] == "CMW":
                CB_list.append(r)

    PS_list2.append("NO Supply for PA")

    return CB_list, PS_list1, PS_list2, CP_list


def PA_Current_Measure(E3632a_1, E3632a_2):
    Sy_current = int(float(E3632a_1.query("MEAS:CURR?")) * 1000)

    if E3632a_2 == "NO Supply for PA":
        Pa_current = 0
    else:
        Pa_current = int(float(E3632a_2.query("MEAS:CURR?")) * 1000)

    return Sy_current, Pa_current


def Ask_query(Callbox, text_area, word_string, query_string):
    response = Callbox.query(f"{query_string}")
    text_area.insert(tk.END, f"{word_string:<40}\t|\t{response}")
    text_area.see(tk.END)


def Check_OPC(Callbox):
    response = Callbox.query("*OPC?")


def update_plot(
    canvas,
    fig,
    ax1,
    ax2,
    ax3,
    rat,
    path,
    Testband,
    channel,
    bandwidth,
    list_Power,
    Power_delta,
    list_Pa_current,
    list_ACLR_max,
):
    if rat == "LTE":
        band_ind = "B"
    elif rat == "NR":
        band_ind = "n"

    ax1.clear()
    x_min, x_max = axis_min_max(list_Power)
    ax1.axis(xmin=x_min, xmax=x_max)
    ax1.set_title(
        f"{rat} {band_ind}{Testband} {path} {channel}CH {bandwidth}MHz PA Current",
        fontsize=8,
    )
    ax1.set_xlabel("Measured Power (dBm)", fontsize=8)
    ax1.set_ylabel("PA Current (mA)", fontsize=8)
    ax1.grid(True, color="black", alpha=0.3, linestyle="--")
    ax1.plot(
        list_Power,
        list_Pa_current,
        marker="o",
        markersize=3,
        color="red",
        linewidth=0.7,
    )
    fig.tight_layout()

    ax2.clear()
    x_min, x_max = axis_min_max(list_Power)
    ax2.axis(xmin=x_min, xmax=x_max)
    ax2.axis(ymin=-1, ymax=2)
    ax2.set_title(f"Power Diff. ( Measured - Target )", fontsize=8)
    ax2.set_xlabel("Measured Power (dBm)", fontsize=8)
    # ax2.set_ylabel("Power Diff (dB)", fontsize=8)
    ax2.grid(True, color="black", alpha=0.3, linestyle="--")
    ax2.plot(list_Power, Power_delta, marker="o", markersize=3, linewidth=0.7)
    fig.tight_layout()

    ax3.clear()
    x_min, x_max = axis_min_max(list_Power)
    ax3.axis(xmin=x_min, xmax=x_max)
    ax3.set_title(f"ACLR (dBc)", fontsize=8)
    ax3.set_xlabel("Measured Power (dBm)", fontsize=8)
    # ax3.set_ylabel("ACLR (dBc)", fontsize=8)
    ax3.grid(True, color="black", alpha=0.3, linestyle="--")
    ax3.plot(list_Power, list_ACLR_max, marker="o", markersize=3, linewidth=0.7)
    ax3.axhline(
        y=-38, linestyle="dashdot", color="red", label="Target"
    )  # Spec 기준선 Drwaing
    fig.tight_layout()

    if len(list_Power) > 5:
        np_PW = np.array(list_Power)
        np_PA = np.array(list_Pa_current)

        z = np.polyfit(np_PW, np_PA, 5)  # (X,Y,차원) 정의
        p = np.poly1d(z)  # 1차원 다항식에 대한 연산을 캡슐화

        ax1.plot(np_PW, p(np_PW), "--", label=f"Trend-line {channel}CH", linewidth=1)
        ax1.legend(fontsize=8, frameon=False, loc="lower right", ncol=6)

    fig.tight_layout()

    canvas.draw()


def axis_min_max(x_list):
    if len(x_list) == 1:
        x_min = 0
        x_max = 25
    else:
        x_min = min(x_list) - 1
        x_max = max(x_list) + 1

    return x_min, x_max

def images2PdfFile(save_dir, filename):
    f_name = f"{os.path.splitext(filename)[0]}.pdf"

    filelist = sorted([file for file in os.listdir(save_dir) if file.endswith(r".jpg")])
    img_files = [os.path.join(save_dir, nm) for nm in filelist]

    with open(f_name, "wb") as pdf_file:
        pdf_file.write(img2pdf.convert(img_files, log_output=None))

    for f in img_files:
        os.remove(f)


def Check_nr_status(Callbox):
    # Meas rdy check
    word = "FETC:NRS:MEAS:MEV:STAT?"
    while True:
        response = Callbox.query(f"{word}")
        time.sleep(0.1)
        if response == "RDY\n":
            break


def Query_aclr(rat, Callbox):
    if rat == "LTE":
        Callbox.write("INIT:LTE:MEAS:MEV")
        Check_OPC(Callbox)
        aclr = Callbox.query("FETCh:LTE:MEASurement:MEValuation:ACLR:AVER?").split(",")
    elif rat == "NR":
        Callbox.write("INIT:NRS:MEAS:MEV")
        Check_OPC(Callbox)
        aclr = Callbox.query("FETCh:NRS:MEASurement:MEValuation:ACLR:AVER?").split(",")
    Lutra2 = round(float(aclr[1]), 2)
    Lutra1 = round(float(aclr[2]), 2)
    Leutra = round(float(aclr[3]), 2)
    TX_power = round(float(aclr[4]), 2)
    Reutra = round(float(aclr[5]), 2)
    Rutra1 = round(float(aclr[6]), 2)
    Rutra2 = round(float(aclr[7]), 2)

    return TX_power, Lutra2, Lutra1, Leutra, Reutra, Rutra1, Rutra2


def define_BW_number(rat, bandwidth):
    if rat == "LTE":
        match bandwidth:
            case 5:
                number = 2
            case 10:
                number = 3
            case 15:
                number = 4
            case 20:
                number = 5

    elif rat == "NR":
        match bandwidth:
            case 5:
                number = 0
            case 10:
                number = 1
            case 15:
                number = 2
            case 20:
                number = 3
            case 25:
                number = 4
            case 30:
                number = 5
            case 40:
                number = 6
            case 50:
                number = 7
            case 60:
                number = 8
            case 80:
                number = 9
            case 90:
                number = 10
            case 100:
                number = 11
            case 70:
                number = 12

    return number


def Set_limit_mask(Callbox, rat, bandwidth):
    if rat == "LTE":
        match bandwidth:
            case 5:
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN50 ON,0MHz,1MHz,-15,K030"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN50 ON,1MHz,2.5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN50 ON,2.5MHz,2.8MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN50 ON,2.8MHz,5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN50 ON,5MHz,6MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN50 ON,6MHz,10MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN50 OFF,10MHz,10MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN50 OFF,10MHz,10MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN50 OFF,10MHz,10MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN50 OFF,10MHz,10MHz,-25,M1"
                )
            case 10:
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN100 ON,0MHz,1MHz,-18,K030"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN100 ON,1MHz,2.5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN100 ON,2.5MHz,2.8MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN100 ON,2.8MHz,5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN100 ON,5MHz,6MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN100 ON,6MHz,10MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN100 ON,10MHz,15MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN100 OFF,15MHz,15MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN100 OFF,15MHz,15MHz,-25,M1"
                )
            case 15:
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN150 ON,0MHz,1MHz,-20,K030"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN150 ON,1MHz,2.5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN150 ON,2.5MHz,2.8MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN150 ON,2.8MHz,5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN150 ON,5MHz,6MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN150 ON,6MHz,10MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN150 ON,10MHz,15MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN150 ON,15MHz,20MHz,-25,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN150 OFF,20MHz,20MHz,-25,M1"
                )
            case 20:
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN200 ON,0,1MHz,-21,K030"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN200 ON,1MHz,2.5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN200 ON,2.5MHz,2.8MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN200 ON,2.8MHz,5MHz,-10,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN200 ON,5MHz,6MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN200 ON,6MHz,10MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN200 ON,10MHz,15MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN200 ON,15MHz,20MHz,-13,M1"
                )
                Callbox.write(
                    "CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN200 ON,20MHz,25MHz,-25,M1"
                )
    elif rat == "NR":
        match bandwidth:
            case 5:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN5 ON,0.015MHz,0.0985MHz,-13.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN5 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN5 ON,5.5MHz,4.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN5 ON,5.5MHz,9.5MHz,-23.5,M1"
                )
            case 10:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN10 ON,0.015MHz,0.0985MHz,-16.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN10 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN10 ON,5.5MHz,9.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN10 ON,10.5MHz,14.5MHz,-23.5,M1"
                )
            case 15:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN15 ON,0.015MHz,0.0985MHz,-18.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN15 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN15 ON,5.5MHz,14.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN15 ON,15.5MHz,19.5MHz,-23.5,M1"
                )
            case 20:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN20 ON,0.015MHz,0.0985MHz,-19.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN20 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN20 ON,5.5MHz,19.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN20 ON,20.5MHz,24.5MHz,-23.5,M1"
                )
            case 25:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN25 ON,0.015MHz,0.0985MHz,-20.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN25 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN25 ON,5.5MHz,24.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN25 ON,25.5MHz,29.5MHz,-23.5,M1"
                )
            case 30:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN30 ON,0.015MHz,0.0985MHz,-21.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN30 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN30 ON,5.5MHz,29.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN30 ON,30.5MHz,34.5MHz,-23.5,M1"
                )
            case 40:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN40 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN40 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN40 ON,5.5MHz,39.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN40 ON,40.5MHz,44.5MHz,-23.5,M1"
                )
            case 50:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN50 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN50 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN50 ON,5.5MHz,49.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN50 ON,50.5MHz,54.5MHz,-23.5,M1"
                )
            case 60:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN60 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN60 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN60 ON,5.5MHz,59.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN60 ON,60.5MHz,64.5MHz,-23.5,M1"
                )
            case 70:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN70 ON,0.015MHz,0.0985MHz,26.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN70 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN70 ON,5.5MHz,69.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN70 ON,70.5MHz,74.5MHz,-23.5,M1"
                )
            case 80:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN80 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN80 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN80 ON,5.5MHz,79.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN80 ON,80.5MHz,84.5MHz,-23.5,M1"
                )
            case 90:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN90 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN90 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN90 ON,5.5MHz,89.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN90 ON,90.5MHz,94.5MHz,-23.5,M1"
                )
            case 100:
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN100 ON,0.015MHz,0.0985MHz,-22.5,K030"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN100 ON,1.5MHz,4.5MHz,-8.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN100 ON,5.5MHz,99.5MHz,-11.5,M1"
                )
                Callbox.write(
                    f"CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN100 ON,100.5MHz,104.5MHz,-23.5,M1"
                )


def Nonsig_lte_tx_measure_setting(
    dut,
    Callbox,
    Equip,
    path_number,
    band,
    bandwidth,
    BW_number,
    NRB,
    PRB,
    rxfreq,
    txfreq,
    text_area,
):
    Equipment = Callbox.query("*IDN?").split(",")[1]
    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        Callbox.write(f"ROUT:GPRF:GEN:SCEN:SAL R118, TX1")
        Callbox.write(
            f"CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, OFF, OFF, OFF, OFF, OFF, OFF, OFF"
        )
        Check_OPC(Callbox)
    elif (Equipment == "CMW") & (Equip == "GPIB0::20::INSTR"):
        Callbox.write("*CLS")
        Callbox.write(f"ROUT:GPRF:GEN:SCEN:SAL RFAC, TX1")
        Check_OPC(Callbox)

    Callbox.write(f"SOUR:GPRF:GEN1:LIST OFF")
    Check_OPC(Callbox)
    Callbox.write(f"SOUR:GPRF:GEN1:RFS:EATT 0.000000")
    Check_OPC(Callbox)
    Callbox.write(f"SOUR:GPRF:GEN1:BBM ARB")
    Check_OPC(Callbox)

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        if band in [38, 39, 40, 41]:
            if bandwidth == 10:
                Callbox.write(
                    f"SOUR:GPRF:GEN1:ARB:FILE"
                    f" 'C:\CMW100_WV\SMU_NodeB_Ant0_LTE_SENS_{bandwidth:02d}MHz_TDD_CFG1_SF_CFG4_SIMO_woCIF_AGL8_RC.wv'"
                )
                Check_OPC(Callbox)
            else:
                Callbox.write(
                    f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_Ant0_TDD_FULL_{BW_number+1:02d}.wv'"
                )
                Check_OPC(Callbox)
        else:
            Callbox.write(
                f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_Ant0_FRC_{bandwidth:02d}MHz.wv'"
            )
            Check_OPC(Callbox)

    elif (Equipment == "CMW") & (Equip == "GPIB0::20::INSTR"):
        if band in [38, 39, 40, 41]:
            if bandwidth == 10:
                Callbox.write(
                    f"SOUR:GPRF:GEN1:ARB:FILE 'D:\SMU_NodeB_Ant0_LTE_SENS_{bandwidth:02d}MHz_TDD_CFG1_SF_CFG4_SIMO_woCIF_AGL8_RC.wv'"
                )
                Check_OPC(Callbox)
            else:
                Callbox.write(
                    f"SOUR:GPRF:GEN1:ARB:FILE 'D:\SMU_NodeB_Ant0_TDD_FULL_{BW_number+1:02d}.wv'"
                )
                Check_OPC(Callbox)
        else:
            Callbox.write(
                f"SOUR:GPRF:GEN1:ARB:FILE 'D:\SMU_NodeB_Ant0_FRC_{bandwidth:02d}MHz.wv'"
            )
            Check_OPC(Callbox)

    word = "SOUR:GPRF:GEN1:ARB:FILE?"
    Ask_query(Callbox, text_area, word, word)

    Callbox.write(f"SOUR:GPRF:GEN1:RFS:FREQ {rxfreq}KHz")
    Callbox.write(f"SOUR:GPRF:GEN1:RFS:LEV -70.000000")

    # Sig gen on
    Callbox.write("SOUR:GPRF:GEN1:STAT ON")
    Check_OPC(Callbox)

    word = "SOUR:GPRF:GEN1:STAT?"  # sig gen on check
    while True:
        response = Callbox.query(f"{word}")
        time.sleep(0.1)
        if response == "ON\n":
            break
    Ask_query(Callbox, text_area, word, word)
    # RX Sync는 Main / Sub 상관없이 동작
    # AT+LSYNC = Main, RX Path, Frequency
    # Main : 0, Main : 0, Frequency
    # CA#1 : 1,  4RX : 1, Frequency
    # CA#2 : 2
    # CA#3 : 3
    response = dut.at_write(f"AT+LSYNC=0,0,{rxfreq}")
    text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
    text_area.see(tk.END)

    response = dut.at_write(
        f"AT+LTXSENDREQ={path_number},{BW_number},{txfreq},{PRB},0,0,2,1,23"
    )

    text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
    text_area.see(tk.END)

    Callbox.write("*CLS")

    if band in [38, 39, 40, 41]:
        Callbox.write("CONF:LTE:MEAS:DMODe TDD")
        Check_OPC(Callbox)
    else:
        Callbox.write("CONF:LTE:MEAS:DMODe FDD")
        Check_OPC(Callbox)
    # time.sleep(0.1)
    Callbox.write(f"CONF:LTE:MEAS:BAND OB{band}")
    Callbox.write(f"CONF:LTE:MEAS:RFS:FREQ {txfreq}KHz")
    Check_OPC(Callbox)
    Callbox.write(f"CONF:LTE:MEAS:MEV:CBAN B{bandwidth*10:03d}")
    Callbox.write("CONF:LTE:MEAS:MEV:MOD:MSCH QPSK")
    Callbox.write(f"CONF:LTE:MEAS:MEV:RBAL:NRB {PRB}")
    Callbox.write("CONF:LTE:MEAS:MEV:RBAL:ORB 0")
    Callbox.write("CONF:LTE:MEAS:MEV:CPR NORM")
    Callbox.write("CONF:LTE:MEAS:MEV:PLC 0")
    Callbox.write("CONF:LTE:MEAS:MEV:DSSP 0")
    Callbox.write("CONF:LTE:MEAS:MEV:RBAL:AUTO OFF")
    Callbox.write("CONF:LTE:MEAS:MEV:MOEX ON")

    Set_limit_mask(Callbox, "LTE", bandwidth)

    Callbox.write("CONFigure:LTE:MEAS:MEValuation:MSLot ALL")
    Callbox.write("CONF:LTE:MEAS:RFS:UMAR 10.000000")
    Callbox.write("CONF:LTE:MEAS:RFS:ENP 28.00")

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        Callbox.write("ROUT:LTE:MEAS:SCEN:SAL R11, RX1")  ##
    elif (Equipment == "CMW") & (Equip == "GPIB0::20::INSTR"):
        Callbox.write("ROUT:LTE:MEAS:SCEN:SAL RFAC, RX1")

    Callbox.write("CONF:LTE:MEAS:RFS:UMAR 10.000000")  # User Margin
    Callbox.write("CONF:LTE:MEAS:MEV:RBAL:AUTO ON")
    Callbox.write("CONF:LTE:MEAS:MEV:SCO:MOD 5")

    Callbox.write("CONF:LTE:MEAS:MEV:SCO:SPEC:ACLR 5")
    Callbox.write("CONF:LTE:MEAS:MEV:SCO:SPEC:SEM 5")

    Callbox.write("TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
    Callbox.write("TRIG:LTE:MEAS:MEV:THR -20.0")
    Callbox.write("CONF:LTE:MEAS:MEV:REP SING")
    # EVM, MagnitudeError, PhaseError, InbandEmissions, EVMversusC, IQ, EquSpecFlatness, TXMeasurement, SpecEmMask, ACLR, RBAllocTable, PowerMonitor, BLER, PowerDynamics
    Callbox.write(
        "CONF:LTE:MEAS:MEV:RES:ALL OFF, OFF, OFF, OFF, OFF, OFF, OFF, ON, OFF, ON, OFF, OFF, OFF, OFF"
    )
    Check_OPC(Callbox)
    Callbox.write("CONF:LTE:MEAS:MEV:MSUB 2, 10, 0")
    Callbox.write("CONF:LTE:MEAS:SCEN:ACT SAL")
    Callbox.write("CONF:LTE:MEAS:RFS:EATT 0.000000")
    Check_OPC(Callbox)

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        Callbox.write("ROUT:GPRF:MEAS:SCEN:SAL R11, RX1")  ##
        Callbox.write("ROUT:LTE:MEAS:SCEN:SAL R11, RX1")  ##
    elif (Equipment == "CMW") & (Equip == "GPIB0::20::INSTR"):
        Callbox.write("ROUT:GPRF:MEAS:SCEN:SAL RFAC, RX1")
        Callbox.write("ROUT:GPRF:GEN:SCEN:SAL RFAC, TX1")
        Callbox.write("CONF:GPRF:MEAS:POW:SCO 2")
        Callbox.write("CONF:GPRF:MEAS:POW:REP SING")
        Callbox.write("CONF:GPRF:MEAS:POW:LIST OFF")
        Callbox.write("TRIGger:GPRF:MEAS:POWer:SOURce 'Free Run'")
        Callbox.write("CONF:GPRF:MEAS:POW:TRIG:SLOP REDG")
        Callbox.write("CONF:GPRF:MEAS:POW:SLEN 5.0e-3")
        Callbox.write("CONF:GPRF:MEAS:POW:MLEN 8.0e-4")
        Callbox.write("TRIGger:GPRF:MEAS:POWer:OFFSet 2.1E-3")
        Callbox.write("TRIG:GPRF:MEAS:POW:MODE ONCE")


def Nonsig_nr_tx_measure_setting(
    dut,
    Callbox,
    Equip,
    path_number,
    Testband,
    bandwidth,
    BW_number,
    PRB,
    infull_offset,
    rxfreq,
    txfreq,
    text_area,
):
    Equipment = Callbox.query("*IDN?").split(",")[1]

    Callbox.write(f"ROUT:GPRF:GEN:SCEN:SAL R118, TX1")
    Callbox.write(
        f"CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, OFF, OFF, OFF, OFF, OFF, OFF, OFF"
    )
    Check_OPC(Callbox)
    Callbox.write(f"SOUR:GPRF:GEN1:LIST OFF")
    Check_OPC(Callbox)
    Callbox.write(f"SOUR:GPRF:GEN1:RFS:EATT 0.000000")
    Check_OPC(Callbox)
    Callbox.write(f"SOUR:GPRF:GEN1:BBM ARB")
    Check_OPC(Callbox)
    Callbox.write(f"CONFigure:NRSub:MEAS:ULDL:PERiodicity MS10")
    Check_OPC(Callbox)

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        if Testband in [38, 40, 41, 77, 78]:
            Callbox.write(
                "SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_NR_10MHz_SCS30_TDD_Sens_MCS0_rescale.wv'"
            )
            Check_OPC(Callbox)
        else:
            Callbox.write(
                "SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_LTE_NR_10MHz_SCS15_FDD_Sens_MCS_0.wv'"
            )
            Check_OPC(Callbox)

    word = "SOUR:GPRF:GEN1:ARB:FILE?"
    Ask_query(Callbox, text_area, word, word)

    Callbox.write(f"SOUR:GPRF:GEN1:RFS:FREQ {rxfreq}KHz")
    Callbox.write(f"SOUR:GPRF:GEN1:RFS:LEV -70.000000")

    # Sig gen on
    Callbox.write("SOUR:GPRF:GEN1:STAT ON")
    Check_OPC(Callbox)

    word = "SOUR:GPRF:GEN1:STAT?"  # sig gen on check
    while True:
        response = Callbox.query(f"{word}")
        time.sleep(0.1)
        if response == "ON\n":
            break
    Ask_query(Callbox, text_area, word, word)

    # AT+NRFSYNC = Main, RX Path, SCS, BW, 256QAM, Frequency
    # Main : 0, Main : 0, SCS 15 : 0,  5M : 0, QPSK : 0, Frequency
    # CA#1 : 1,  4RX : 1, SCS 30 : 1, 10M : 1, 256Q : 3, Frequency
    # CA#2 : 2,  6RX : 2,
    # CA#3 : 3

    # AT+NTXSENDREQ=TX Path, Frequency, BW, SCS, RBSize, RBOffset, MCS, Waveform type, TX Level
    # Main : 0, Frequency,  5M : 0, SCS 15 : 0, RBSize, RBOffset, QPSK : 2, DFT-S : 0, TX Level
    #  Sub : 1, Frequency, 10M : 1, SCS 30 : 1, RBSize, RBOffset, 256Q : 8,    CP : 1, TX Level
    # MIMO : 20
    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        if Testband in [38, 40, 41, 77, 78]:  # SCS 30
            response = dut.at_write(f"AT+NRFSYNC=0,0,1,{BW_number},0,{rxfreq}")
            text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
            text_area.see(tk.END)
            response = dut.at_write(
                f"AT+NTXSENDREQ={path_number},{txfreq},{BW_number},1,{PRB},{infull_offset},2,0,23"
            )
            text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
            text_area.see(tk.END)
            Callbox.write("CONF:NRS:MEAS:MEV:DMODe TDD")
            Check_OPC(Callbox)
            scs = 30
        else:  # SCS 15
            response = dut.at_write(f"AT+NRFSYNC=0,0,0,{BW_number},0,{rxfreq}")
            text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
            text_area.see(tk.END)
            response = dut.at_write(
                f"AT+NTXSENDREQ={path_number},{txfreq},{BW_number},0,{PRB},{infull_offset},2,0,23"
            )
            text_area.insert(tk.END, f"{response[0]:<40}\t|\t{response[2::2]}\n")
            text_area.see(tk.END)
            Callbox.write("CONF:NRS:MEAS:MEV:DMODe FDD")
            Check_OPC(Callbox)
            scs = 15
    # time.sleep(0.1)
    Callbox.write(f"CONF:NRS:MEAS:BAND OB{Testband}")
    Callbox.write(f"CONF:NRS:MEAS:RFS:FREQ {txfreq}KHz")
    Check_OPC(Callbox)
    Callbox.write("CONF:NRS:MEAS:MEV:PLC 0")
    Callbox.write("CONF:NRS:MEAS:MEV:MOEX ON")
    Callbox.write(f"CONF:NRS:MEAS:MEV:BWC S{scs}K, B{bandwidth:03d}")

    Set_limit_mask(Callbox, "NR", bandwidth)

    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:DFTPrecoding ON")

    if (Equipment == "CMW") & (Equip == "TCPIP::127.0.0.1::INSTR"):
        if Testband in [38, 40, 41, 77, 78]:  # SCS 30
            Callbox.write(
                f"CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig QPSK,A,OFF,{PRB},{infull_offset},14,0,T1,SING,0,2"
            )
        else:
            Callbox.write(
                f"CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig QPSK,A,OFF,{PRB},{infull_offset},14,0,T1,SING,0,2"
            )

    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:PCOMp OFF, 6000E+6")
    Check_OPC(Callbox)
    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:REPetition SING")
    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:PLCid 0")
    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:CTYPe PUSC")
    Callbox.write("CONF:NRS:MEAS:ULDL:PER MS25")
    Callbox.write(f"CONF:NRS:MEAS:ULDL:PATT S{scs}K, 3,0,1,14")
    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL")
    Callbox.write("ROUT:NRS:MEAS:SCEN:SAL R11, RX1")
    Callbox.write("CONF:NRS:MEAS:RFS:ENP 28.00")
    Callbox.write("CONF:NRS:MEAS:RFS:UMAR 10.000000")
    Callbox.write("CONF:NRS:MEAS:MEV:SCO:MOD 5")
    Callbox.write("CONF:NRS:MEAS:MEV:SCO:SPEC:ACLR 5")
    Callbox.write("CONF:NRS:MEAS:MEV:SCO:SPEC:SEM 5")
    Callbox.write("TRIG:NRS:MEAS:MEV:SOUR 'GPRF GEN1: Restart Marker'")
    Callbox.write("TRIG:NRS:MEAS:MEV:THR -20.0")
    Callbox.write("CONF:NRS:MEAS:MEV:REP SING")
    # EVM, MagnitudeError, PhaseError, InbandEmissions, EVMversusC, IQ, EquSpecFlatness, TXMeasurement, SpecEmMask, ACLR, PowerMonitor, PowerDynamics, TxPower
    Callbox.write(
        "CONF:NRS:MEAS:MEV:RES:ALL OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF, OFF"
    )
    Callbox.write("CONF:NRS:MEAS:MEV:RES:ACLR ON")
    Callbox.write("CONF:NRS:MEAS:MEV:NSUB 10")
    Callbox.write("CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL")
    Callbox.write("CONF:NRS:MEAS:SCEN:ACT SAL")
    Callbox.write("CONF:NRS:MEAS:RFS:EATT 0.000000")
    Check_OPC(Callbox)
    Callbox.write("ROUT:GPRF:MEAS:SCEN:SAL R11, RX1")
    Check_OPC(Callbox)
    Callbox.write("ROUT:NRS:MEAS:SCEN:SAL R11, RX1")
    Check_OPC(Callbox)
    Callbox.write("INIT:NRS:MEAS:MEV")
    Check_OPC(Callbox)